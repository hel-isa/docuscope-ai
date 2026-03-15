from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def parse_xlsx(file_path: str | Path) -> dict:
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"XLSX file not found: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    lines: list[str] = []
    tables_detected = False

    for ws in wb.worksheets:
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            values = [str(v) for v in row if v is not None and str(v).strip()]
            if values:
                row_count += 1
                lines.append(" | ".join(values))
            if row_count >= 200:
                break
        if row_count > 1:
            tables_detected = True

    props = getattr(wb, "properties", None)
    embedded_metadata = {
        "creator": getattr(props, "creator", None),
        "title": getattr(props, "title", None),
        "subject": getattr(props, "subject", None),
        "description": getattr(props, "description", None),
        "keywords": getattr(props, "keywords", None),
    }

    return {
        "text": "\n".join(lines),
        "page_count": len(wb.worksheets),
        "author_safe": embedded_metadata.get("creator"),
        "embedded_metadata": {k: v for k, v in embedded_metadata.items() if v},
        "ocr_needed": False,
        "ocr_used": False,
        "tables_detected": tables_detected,
        "signatures_detected": False,
        "stamps_detected": False,
    }
